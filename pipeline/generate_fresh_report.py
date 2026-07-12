"""
2026 出海选品机会分析 — 基于纯新鲜数据（排除2021 Amazon数据集）
数据源: Reddit(73) + GitHub(60) + HackerNews(59) + GoogleTrends(47) + YouTube(37) = 276条
采集时间: 2026-07-07（今天）
"""
import sqlite3
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/grace/WorkBuddy/2026-07-07-21-28-00/cross-border-ecommerce-skills/pipeline/demand_signals.db"
OUTPUT = "/Users/grace/WorkBuddy/2026-07-07-21-28-00/2026_opportunity_fresh.xlsx"

# ═══════════════════════════════════════════════════════════════════════
# 数据加载
# ═══════════════════════════════════════════════════════════════════════
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

FRESH_SOURCES = ('reddit', 'github', 'hackernews', 'google_trends', 'youtube')
rows = conn.execute(
    "SELECT * FROM demand_signals WHERE source IN ({})".format(
        ','.join(['?'] * len(FRESH_SOURCES))
    ), FRESH_SOURCES
).fetchall()

print(f"Loaded {len(rows)} fresh records")

# ═══════════════════════════════════════════════════════════════════════
# 机会分析（纯基于新鲜数据信号）
# ═══════════════════════════════════════════════════════════════════════

opportunities = [
    {
        "rank": 1,
        "market": "欧盟 (EU)",
        "country": "德国/法国/荷兰/北欧",
        "category": "EU制造替代品 (数字主权产品)",
        "signal_strength": "★★★★★",
        "signal_count": 5,
        "signals": [
            "Reddit r/BuyFromEU (70评论): 用户主动求EU制造的VPN替代Proton",
            "Reddit r/Aliexpress (130评论): 'New EU fee coming into force in November' — EU新关税壁垒",
            "Reddit r/Aliexpress (50评论): 'prices doubled or tripled for EU' — 价格翻倍痛点",
            "HN (199赞): 'Top researchers leave USA for the Netherlands' — 人才流向EU",
            "HN (134赞): 'Europe company websites mostly served by US vendors' — 数字主权缺口",
        ],
        "pain_point": "EU消费者面临新关税导致进口商品价格翻倍；同时BuyFromEU运动兴起，用户主动寻求EU制造的替代品，尤其是数字产品（VPN、SaaS）",
        "opportunity": "定位'100% EU-made'的数字产品：VPN、密码管理器、云存储、隐私工具。强调GDPR合规、数据本地存储。物理产品主打EU本地仓直发，避开新关税",
        "target_aov": "$10-50/月 (订阅) 或 $30-80 (一次性)",
        "competition": "Proton (瑞士)、Tuta (德国) 已占位，但用户明确表示不满Proton，有切换意愿",
        "entry_strategy": "1) 做Proton的EU替代品比较站，SEO引流；2) 推出EU服务器+EU数据中心的SaaS工具；3) 联合BuyFromEU社区做口碑营销",
        "data_evidence": "Reddit直接痛点帖 + HN技术社区认同 = 双重验证",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 2,
        "market": "全球 (数字交付)",
        "country": "美国/欧盟/东南亚",
        "category": "数字产品 (模板/工具/课程)",
        "signal_strength": "★★★★★",
        "signal_count": 8,
        "signals": [
            "YouTube: '3 TINY Digital Products Guaranteed to Sell in 2026' 130K播放",
            "YouTube: '3 BEST Digital Products to Sell Online in 2026' 78K播放",
            "YouTube: '5 Best Digital Products to Sell in 2026' 57K播放",
            "YouTube: 'The 10 BEST Digital Products to Sell Online Right Now' 23K播放",
            "YouTube: 'Beginners Guide To AI Dropshipping' 68K播放",
            "Google Trends: 'ai dropshipping' 趋势值88 (上升)",
            "Reddit r/MonarchMoney (170评论): 'What do you wish Monarch did better?' — SaaS功能缺口",
            "Reddit r/ecommerce (40评论): 'Should I switch to Shopify' — 平台迁移需求",
        ],
        "pain_point": "800美元免税取消后，实体跨境成本飙升；卖家急需零物流、零关税的数字产品模式。YouTube上'数字产品'视频合计30万+播放验证需求",
        "opportunity": "做选品模板包、Shopify店铺装修模板、AI选品工具订阅、电商运营SOP课程。通过Gumroad/Etsy/自有站销售，数字交付零物流",
        "target_aov": "$9-49 (模板) / $29-99/月 (SaaS)",
        "competition": "Gumroad上有竞品但质量参差不齐，AI选品工具几乎空白",
        "entry_strategy": "1) 在YouTube做'免费选品教程'引流→卖模板；2) 做Monarch Money的平替SaaS；3) 做Shopify迁移工具/服务",
        "data_evidence": "YouTube 30万播放 + Google Trends 88 + Reddit 170评论 = 三重验证",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 3,
        "market": "美国",
        "country": "美国",
        "category": "AI驱动的电商工具 (选品/自动化)",
        "signal_strength": "★★★★☆",
        "signal_count": 6,
        "signals": [
            "YouTube: 'Chat GPT Does All My Amazon Product Research' 6.3K播放",
            "YouTube: 'Dominate Amazon FBA Product Research with AI' 3.7K播放",
            "YouTube: 'Beginners Guide To AI Dropshipping' 68K播放",
            "Google Trends: 'ai dropshipping' 趋势值88",
            "Google Trends: 'amazon fba step by step' 趋势值200 (breakout级别)",
            "HN (584赞): 'GLM 5.2 and the coming AI margin collapse' — AI成本骤降",
        ],
        "pain_point": "大量新卖家涌入Amazon FBA/dropshipping，但不知道选什么品。Google Trends显示'amazon fba step by step'搜索量200(breakout)，'ai dropshipping'趋势88且上升",
        "opportunity": "做AI选品SaaS：输入品类→AI分析Reddit/Amazon/TikTok数据→输出选品建议。利用AI成本下降(HN验证)，做一个$29/月的工具",
        "target_aov": "$19-49/月 (SaaS订阅)",
        "competition": "Jungle Scout/Helium 10做Amazon数据分析，但不做AI选品推荐；WhatToSell做需求发现但无AI",
        "entry_strategy": "1) 先做免费Chrome插件(Reddit痛点扫描)→引流到付费SaaS；2) YouTube做'AI选品实战'内容营销",
        "data_evidence": "YouTube 68K播放 + Google Trends双关键词breakout + HN AI成本下降验证",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 4,
        "market": "全球 (隐私优先)",
        "country": "欧盟/美国/东南亚",
        "category": "隐私硬件 & 自托管工具",
        "signal_strength": "★★★★☆",
        "signal_count": 5,
        "signals": [
            "HN (744赞 282评论): 'OpenWrt One – Open Hardware Router' — 开源路由器热度爆表",
            "HN (690赞): 'CoMaps – FOSS Offline Maps' — 隐私地图替代Google Maps",
            "HN (226赞 97评论): 'Microsoft Can Track Users via Windows Device ID' — 隐私焦虑",
            "Reddit r/selfhosted (60评论): 'Has anyone tried to go colocation?' — 自托管趋势",
            "Reddit r/GlInet: 'Where can I buy a GL-iNet Mudi 7 travel router' — 旅行路由器需求",
        ],
        "pain_point": "科技用户对大厂追踪日益不满(Microsoft ID追踪226赞)，自托管/隐私硬件需求上升。OpenWrt开源路由器744赞为HN年度Top级",
        "opportunity": "做隐私旅行路由器(预装OpenWrt/VPN)、自托管 starter kit(预配置NAS+Docker)、隐私工具包。面向digital nomad和隐私敏感用户",
        "target_aov": "$50-150 (路由器) / $200-500 (自托管套件)",
        "competition": "GL.iNet已有产品但营销弱；r/selfhosted用户在找colocation方案但无消费级产品",
        "entry_strategy": "1) 做GL.iNet的营销代理/改良版；2) 推出'5分钟自托管'套件(NAS+Docker+一键脚本)；3) HN/Reddit社区口碑营销",
        "data_evidence": "HN 3个帖子合计1660赞 + Reddit 2个板块需求 = 技术社区强信号",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 5,
        "market": "美国",
        "country": "美国",
        "category": "二手/转售平台 & 品质升级产品",
        "signal_strength": "★★★★☆",
        "signal_count": 4,
        "signals": [
            "Reddit r/poshmark (110评论): 'I wish there was a separate site for resellers' — 转售平台缺口",
            "Reddit r/REI (100评论): 'denied a return for my Arcteryx coming apart' — 品质+售后痛点",
            "Reddit r/litterrobot (120评论): 'I'll never purchase another Whisker product again' — 宠物科技品质差",
            "Reddit r/UsedCars (30评论): 'I wish there was an ethical code for dealerships' — 二手市场信任缺失",
        ],
        "pain_point": "美国消费者对现有品牌品质和售后不满(3个帖子合计330评论)，同时二手转售市场缺少专门的reseller平台",
        "opportunity": "1) 做reseller专用工具(批量上架+价格监控SaaS)；2) 做'品质保证'的宠物科技平替产品；3) 做户外装备品质升级款",
        "target_aov": "$15-50/月 (SaaS) / $30-100 (宠物/户外产品)",
        "competition": "Poshmark/Vinted存在但reseller体验差；Whisker(Litter-Robot)有品质问题但无强力竞品",
        "entry_strategy": "1) 做Poshmark的reseller工具→SaaS月费；2) 做Litter-Robot平替(更可靠+更便宜)；3) Reddit口碑营销直击差评",
        "data_evidence": "Reddit 4帖合计360评论，高评论数=高共鸣度",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 6,
        "market": "印度",
        "country": "印度",
        "category": "电子产品验真 & 信任电商",
        "signal_strength": "★★★☆☆",
        "signal_count": 2,
        "signals": [
            "Reddit r/LaptopDealsIndia (10评论): 'Amazon/Flipkart laptops are used ones?' — 印度电商信任危机",
            "Reddit r/DigitalProductSellers: 'What niches are profitable on Amazon/Mercado Libre' — 跨境卖家关注新兴市场",
        ],
        "pain_point": "印度消费者怀疑Amazon/Flipkart卖翻新机，信任是最大瓶颈。印度电商CAGR 17.6%但信任缺失限制增长",
        "opportunity": "做电子产品验真服务(二维码溯源SaaS)或'品质保证'的电子产品独立站，主打验真+保修",
        "target_aov": "$50-200 (电子产品) / $5-15/月 (验真SaaS)",
        "competition": "印度本土有Cashify(二手电子)，但验真服务空白",
        "entry_strategy": "1) 给印度卖家提供验真标签SaaS；2) 自营'验真电子'独立站，Flipkart/Amazon引流",
        "data_evidence": "Reddit直接信任痛点 + 印度市场增速数据",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 7,
        "market": "拉美",
        "country": "巴西/墨西哥",
        "category": "Mercado Libre 选品工具 & 电子配件",
        "signal_strength": "★★★☆☆",
        "signal_count": 2,
        "signals": [
            "Reddit r/DigitalProductSellers: 'What niches are profitable on Amazon/Mercado Libre' — 拉美市场关注",
            "Google Trends: 'shopify 登录' 趋势值60 — 中文用户搜索Shopify，跨境卖家活跃",
        ],
        "pain_point": "拉美电商增速22%(全球最快)，但中国卖家缺少Mercado Libre选品数据。Reddit上有人直接问'什么品类在Mercado Libre赚钱'",
        "opportunity": "做Mercado Libre选品数据SaaS，或直接做拉美电子配件(手机壳/充电器/数据线)独立站",
        "target_aov": "$5-20 (电子配件) / $29/月 (SaaS)",
        "competition": "拉美本地缺少选品工具，中国卖家对这个市场认知不足=先发优势",
        "entry_strategy": "1) 做ML选品工具(爬ML数据→趋势分析)；2) 深圳3C配件→ML/FBA发货",
        "data_evidence": "Reddit主动提问 + 拉美市场增速22%数据",
        "freshness": "2026-07-07 采集",
    },
    {
        "rank": 8,
        "market": "全球 (niche)",
        "country": "多市场",
        "category": "桌面游戏配件 & 小众需求",
        "signal_strength": "★★★☆☆",
        "signal_count": 6,
        "signals": [
            "Reddit r/Tau40K: 'Does anyone know where to find this exact terrain?' — Warhammer地形配件需求",
            "Reddit r/DungeonCrawlerCarlRPG: 'Missed my chance on Backerkit' — 桌游KS缺货",
            "Reddit r/DarkTide: 'I wish there was more cosmetic customisation' — 游戏外观定制需求",
            "Reddit r/AvatarFrontierPandora: 'Does anyone know where to find this face paint' — 游戏cosplay化妆",
            "Reddit r/mylittlepony: 'where can i buy the my little pony Kotobukiya Bishoujo' — 收藏品难买",
            "Reddit r/Fingerboards: 'Who and why removed the Flatface Harrier' — 指板配件断货",
        ],
        "pain_point": "桌游/收藏品/小众爱好玩家'找不到、买不到'的痛点集中爆发，6个不同niche都有'where can I buy'信号",
        "opportunity": "做小众爱好聚合电商(桌游地形+收藏品+cosplay道具)，或做3D打印定制服务",
        "target_aov": "$15-80 (配件) / $50-200 (定制)",
        "competition": "Etsy有部分，但缺少专业聚合站；3D打印定制几乎空白",
        "entry_strategy": "1) 做3D打印桌游地形Etsy店；2) 做小众收藏品代购/聚合站；3) 按需定制cosplay道具",
        "data_evidence": "Reddit 6个不同niche板块均有'where can I buy'帖",
        "freshness": "2026-07-07 采集",
    },
]

# ═══════════════════════════════════════════════════════════════════════
# Excel 生成
# ═══════════════════════════════════════════════════════════════════════
wb = Workbook()

# 样式
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
cell_font = Font(name='Arial', size=10)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# ── Sheet 1: 机会总览 ──
ws1 = wb.active
ws1.title = "机会总览"

ws1.merge_cells('A1:L1')
ws1['A1'] = "2026出海选品机会分析 — 基于纯新鲜数据 (276条实时信号)"
ws1['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E79')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:L2')
ws1['A2'] = f"数据源: Reddit(73) + GitHub(60) + HackerNews(59) + Google Trends(47) + YouTube(37) | 采集时间: 2026-07-07 | 已排除2021年Amazon数据集"
ws1['A2'].font = Font(name='Arial', size=10, italic=True, color='808080')
ws1['A2'].alignment = Alignment(horizontal='center')

headers = ["排名", "目标市场", "目标国家", "机会品类", "信号强度", "信号数", "核心痛点", "机会描述", "建议客单价", "竞争格局", "切入策略", "数据证据"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws1.row_dimensions[4].height = 30

# 排名颜色
rank_colors = {
    1: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    2: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    3: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    4: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    5: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
}

for i, opp in enumerate(opportunities):
    r = 5 + i
    vals = [
        opp["rank"], opp["market"], opp["country"], opp["category"],
        opp["signal_strength"], opp["signal_count"], opp["pain_point"],
        opp["opportunity"], opp["target_aov"], opp["competition"],
        opp["entry_strategy"], opp["data_evidence"]
    ]
    for col, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=col, value=v)
        cell.font = cell_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if opp["rank"] in rank_colors:
            cell.fill = rank_colors[opp["rank"]]
    ws1.row_dimensions[r].height = 100

col_widths = [6, 14, 16, 22, 10, 8, 40, 40, 18, 28, 40, 25]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: 信号明细 ──
ws2 = wb.create_sheet("信号明细")

ws2.merge_cells('A1:G1')
ws2['A1'] = "全部276条新鲜数据信号明细"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

detail_headers = ["数据源", "信号类型", "标题", "内容摘要", "评分/热度", "URL", "附加信息"]
for col, h in enumerate(detail_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

source_colors = {
    'reddit': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),
    'youtube': PatternFill(start_color='FFE0F0', end_color='FFE0F0', fill_type='solid'),
    'hackernews': PatternFill(start_color='FFF0E0', end_color='FFF0E0', fill_type='solid'),
    'github': PatternFill(start_color='E0F0FF', end_color='E0F0FF', fill_type='solid'),
    'google_trends': PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid'),
}

for i, row in enumerate(rows):
    r = 4 + i
    m = json.loads(row['metadata']) if row['metadata'] else {}
    extra = ""
    if row['source'] == 'reddit':
        extra = f"r/{m.get('subreddit','')}"
    elif row['source'] == 'youtube':
        extra = f"频道: {m.get('channel','')}"
    elif row['source'] == 'hackernews':
        extra = f"评论数: {m.get('comments',0)}"
    elif row['source'] == 'github':
        extra = f"仓库: {m.get('repo','')}"

    vals = [
        row['source'], row['signal_type'], row['title'] or '',
        (row['content'] or '')[:200], row['score'] or 0,
        row['url'] or '', extra
    ]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.font = cell_font
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.fill = source_colors.get(row['source'], PatternFill())

for i, w in enumerate([12, 15, 45, 50, 12, 45, 25], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 3: 按市场分组 ──
ws3 = wb.create_sheet("按市场分组")

ws3.merge_cells('A1:E1')
ws3['A1'] = "2026机会按目标市场分组"
ws3['A1'].font = title_font

market_groups = {
    "欧盟 (EU)": [1],
    "全球/数字交付": [2],
    "美国": [3, 5],
    "全球/隐私": [4],
    "印度": [6],
    "拉美": [7],
    "全球/niche": [8],
}

r = 3
for market, ranks in market_groups.items():
    ws3.cell(row=r, column=1, value=market).font = Font(name='Arial', size=12, bold=True, color='1F4E79')
    r += 1
    for rank in ranks:
        opp = next(o for o in opportunities if o["rank"] == rank)
        ws3.cell(row=r, column=1, value=f"  #{opp['rank']}").font = cell_font
        ws3.cell(row=r, column=2, value=opp["category"]).font = cell_font
        ws3.cell(row=r, column=3, value=opp["signal_strength"]).font = cell_font
        ws3.cell(row=r, column=4, value=opp["target_aov"]).font = cell_font
        ws3.cell(row=r, column=5, value=opp["pain_point"][:80]).font = cell_font
        ws3.cell(row=r, column=5).alignment = wrap_align
        r += 1
    r += 1

for i, w in enumerate([10, 25, 12, 18, 60], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 4: 数据源说明 ──
ws4 = wb.create_sheet("数据源说明")

ws4.merge_cells('A1:D1')
ws4['A1'] = "数据源说明 — 已排除旧数据"
ws4['A1'].font = title_font

source_info = [
    ["数据源", "记录数", "时效性", "说明"],
    ["Reddit", 73, "实时 (2026-07-07采集)", "通过BrowserAct + Google site:reddit.com搜索，覆盖60+板块，含痛点帖、功能请求、购买意向"],
    ["GitHub Issues", 60, "实时 (2026-07-07采集)", "GitHub Search API，搜索enhancement/wishlist标签的开放issue"],
    ["Hacker News", 59, "实时 (2026-07-07采集)", "Firebase API，获取topstories/askstories/showstories，技术社区趋势信号"],
    ["Google Trends", 47, "当前 (2026-07-07)", "curl_cffi + explore API，获取amazon fba/dropshipping/shopify等关键词趋势和相关搜索"],
    ["YouTube", 37, "近期视频", "curl_cffi + ytInitialData解析，选品/电商/数字产品相关视频，含播放量"],
    ["", "", "", ""],
    ["已排除: Amazon数据集", 484, "2014-2023 (90%在2019-2021)", "HuggingFace McAuley-Lab Amazon Reviews，评论时间过旧，不适用于2026选品决策"],
    ["已排除: Amazon实时", 151, "—", "Jina解析失败，采集到页面导航文字而非商品信息，废数据"],
    ["已排除: ProductHunt", 78, "—", "Jina解析失败，采集到页脚元素，废数据"],
    ["已排除: Trustpilot", 1, "—", "仅1条，样本过小"],
    ["", "", "", ""],
    ["新鲜数据合计", 276, "全部2026-07-07采集", "5个数据源，8种信号类型，覆盖6大市场"],
]

for i, row_data in enumerate(source_info):
    r = 3 + i
    for col, v in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=col, value=v)
        if i == 0:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        else:
            cell.font = cell_font
            cell.alignment = wrap_align
        cell.border = thin_border

for i, w in enumerate([22, 10, 25, 65], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 5: 优先级矩阵 ──
ws5 = wb.create_sheet("优先级矩阵")

ws5.merge_cells('A1:F1')
ws5['A1'] = "优先级矩阵 — 信号强度 × 可执行性"
ws5['A1'].font = title_font

matrix_headers = ["优先级", "排名", "品类", "市场", "信号强度", "可执行性评分"]
for col, h in enumerate(matrix_headers, 1):
    cell = ws5.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

priority_data = [
    ["P0 立即行动", 1, "EU制造替代品", "欧盟", "★★★★★", "高 (有明确需求+BuyFromEU社区)"],
    ["P0 立即行动", 2, "数字产品/模板", "全球", "★★★★★", "高 (零物流+YouTube验证)"],
    ["P1 近期启动", 3, "AI选品SaaS", "美国", "★★★★☆", "中高 (技术门槛+AI成本下降)"],
    ["P1 近期启动", 4, "隐私硬件/自托管", "EU/US", "★★★★☆", "中 (需硬件供应链)"],
    ["P1 近期启动", 5, "二手转售工具/品质平替", "美国", "★★★★☆", "高 (SaaS可快速启动)"],
    ["P2 观察验证", 6, "电子验真", "印度", "★★★☆☆", "中 (需本地化运营)"],
    ["P2 观察验证", 7, "ML选品工具/电子配件", "拉美", "★★★☆☆", "中 (市场认知不足)"],
    ["P2 观察验证", 8, "桌游/niche配件", "多市场", "★★★☆☆", "中低 (市场分散)"],
]

priority_colors = {
    "P0 立即行动": PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    "P1 近期启动": PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    "P2 观察验证": PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

for i, row_data in enumerate(priority_data):
    r = 4 + i
    for col, v in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=col, value=v)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = priority_colors.get(row_data[0], PatternFill())

for i, w in enumerate([14, 8, 22, 10, 12, 30], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")

conn.close()

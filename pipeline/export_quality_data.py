#!/usr/bin/env python3
"""
Export quality demand signals to Excel.
Excludes noise sources (amazon nav, producthunt nav, trustpilot single record).
Normalizes posted_at timestamps to readable datetime.
"""
import sqlite3
import json
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "demand_signals.db"
OUTPUT_PATH = "/Users/grace/WorkBuddy/2026-07-07-21-28-00/demand_signals_quality.xlsx"

# Quality sources only (exclude amazon=nav noise, producthunt=nav noise, trustpilot=1 record)
QUALITY_SOURCES = [
    "amz_dataset",    # 484 Amazon真实评论
    "reddit",         # 73 Reddit用户痛点
    "github",         # 60 GitHub feature request
    "hackernews",     # 59 HN讨论
    "google_trends",  # 47 趋势信号
    "youtube",        # 37 视频 content
]

SOURCE_CN = {
    "amz_dataset": "Amazon评论数据集",
    "reddit": "Reddit",
    "github": "GitHub Issues",
    "hackernews": "Hacker News",
    "google_trends": "Google Trends",
    "youtube": "YouTube",
}

SIGNAL_TYPE_CN = {
    "pain_point": "痛点",
    "review": "评论",
    "feature_request": "功能需求",
    "discussion": "讨论",
    "trend": "趋势",
    "video_content": "视频内容",
    "product_listing": "产品列表",
    "purchase_intent": "购买意向",
}

SCORE_LABEL_CN = {
    "rating": "评分(1-5星)",
    "upvotes": "点赞数",
    "comments": "评论数",
    "views": "播放量",
    "search_volume": "搜索热度",
    "trend_value": "趋势值",
}


def normalize_posted_at(raw, source, metadata):
    """Convert various timestamp formats to readable datetime string."""
    if not raw or raw == "":
        return ""

    # amz_dataset: Unix milliseconds (string like "1541698828578")
    if source == "amz_dataset":
        try:
            ts = int(raw) / 1000
            return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
        except (ValueError, OSError):
            return raw

    # github: ISO with Z ("2022-06-02T12:01:18Z")
    if source == "github":
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            return raw

    # youtube: relative time ("5mo ago", "1y ago", "3w ago")
    if source == "youtube":
        now = datetime(2026, 7, 7, 23, 0)
        match = re.match(r"(\d+)([dwmMy])\s*ago", raw)
        if match:
            num = int(match.group(1))
            unit = match.group(2)
            if unit == "d":
                delta = timedelta(days=num)
            elif unit == "w":
                delta = timedelta(weeks=num)
            elif unit == "m":
                delta = timedelta(days=num * 30)
            elif unit == "M":
                delta = timedelta(days=num * 30)
            elif unit == "y":
                delta = timedelta(days=num * 365)
            else:
                return raw
            approx = now - delta
            return f"~{approx.strftime('%Y-%m-%d')} (约{raw})"
        return raw

    # reddit/hackernews/google_trends: these are collection timestamps
    # Reddit posted_at is collection time (no real post time stored)
    # HackerNews posted_at is collection time (real time was in HN API but not saved)
    # Google Trends is aggregate data (no single post time)
    if source in ("reddit", "hackernews", "google_trends"):
        try:
            dt = datetime.fromisoformat(raw)
            return dt.strftime("%Y-%m-%d %H:%M") + " (采集时间)"
        except (ValueError, TypeError):
            return raw + " (采集时间)"

    return raw


def extract_metadata_summary(metadata_json, source):
    """Extract human-readable summary from metadata JSON."""
    if not metadata_json or metadata_json == "{}":
        return ""
    try:
        m = json.loads(metadata_json)
    except json.JSONDecodeError:
        return ""

    if source == "amz_dataset":
        parts = []
        if m.get("asin"):
            parts.append(f"ASIN: {m['asin']}")
        if m.get("rating") is not None:
            parts.append(f"评分: {m['rating']}星")
        if m.get("verified_purchase"):
            parts.append("已验证购买")
        if m.get("helpful_votes", 0) > 0:
            parts.append(f"有用票数: {m['helpful_votes']}")
        return " | ".join(parts)

    if source == "reddit":
        parts = []
        if m.get("subreddit"):
            parts.append(f"板块: {m['subreddit']}")
        if m.get("search_query"):
            parts.append(f"搜索: {m['search_query']}")
        return " | ".join(parts)

    if source == "hackernews":
        parts = []
        if m.get("hn_id"):
            parts.append(f"HN ID: {m['hn_id']}")
        if m.get("comments") is not None:
            parts.append(f"评论数: {m['comments']}")
        if m.get("type"):
            parts.append(f"类型: {m['type']}")
        return " | ".join(parts)

    if source == "youtube":
        parts = []
        if m.get("channel"):
            parts.append(f"频道: {m['channel']}")
        if m.get("views") is not None:
            parts.append(f"播放量: {m['views']:,}")
        if m.get("video_id"):
            parts.append(f"视频ID: {m['video_id']}")
        return " | ".join(parts)

    if source == "google_trends":
        parts = []
        if m.get("keyword"):
            parts.append(f"关键词: {m['keyword']}")
        if m.get("parent_keyword"):
            parts.append(f"关联词: {m['parent_keyword']}")
        if m.get("related_query"):
            parts.append(f"相关搜索: {m['related_query']}")
        if m.get("data_points"):
            parts.append(f"数据点: {m['data_points']}")
        if m.get("timeframe"):
            parts.append(f"时间范围: {m['timeframe']}")
        return " | ".join(parts)

    return json.dumps(m, ensure_ascii=False)[:100]


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    ws = wb.active
    ws.title = "质量反馈数据"

    # ── Title row ──
    headers = [
        "序号", "数据源", "信号类型", "标题", "反馈内容",
        "反馈时间", "评分/热度", "评分类型", "URL", "作者",
        "关键词", "品类", "附加信息", "采集时间"
    ]

    # Style: header
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Source color fills
    source_fills = {
        "amz_dataset": PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),
        "reddit": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "github": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "hackernews": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        "google_trends": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "youtube": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    }

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    body_font = Font(name="Microsoft YaHei", size=10)
    body_align = Alignment(vertical="top", wrap_text=True)

    row_idx = 2
    for source in QUALITY_SOURCES:
        rows = conn.execute("""
            SELECT * FROM demand_signals 
            WHERE source = ? 
            ORDER BY score DESC
        """, (source,)).fetchall()

        for r in rows:
            posted_at = normalize_posted_at(r["posted_at"], source, r["metadata"])
            meta_summary = extract_metadata_summary(r["metadata"], source)

            # Parse keywords
            keywords_str = ""
            if r["keywords"]:
                try:
                    kws = json.loads(r["keywords"])
                    keywords_str = ", ".join(kws)
                except json.JSONDecodeError:
                    keywords_str = r["keywords"]

            # Score display
            score_val = r["score"]
            if score_val is not None:
                if source == "youtube" and score_val > 1000:
                    score_display = f"{score_val:,.0f}"
                elif source == "google_trends" and score_val == score_val:
                    score_display = f"{score_val:.0f}"
                elif source == "amz_dataset":
                    score_display = f"{score_val:.1f}"
                else:
                    score_display = str(int(score_val)) if score_val == int(score_val) else str(score_val)
            else:
                score_display = ""

            score_label_display = SCORE_LABEL_CN.get(r["score_label"], r["score_label"] or "")

            row_data = [
                row_idx - 1,
                f"{SOURCE_CN.get(source, source)} ({source})",
                SIGNAL_TYPE_CN.get(r["signal_type"], r["signal_type"]),
                r["title"] or "",
                r["content"] or "",
                posted_at,
                score_display,
                score_label_display,
                r["url"] or "",
                r["author"] or "",
                keywords_str,
                r["category"] or "",
                meta_summary,
                r["collected_at"] or "",
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border
                # Source-based row coloring
                if col <= 2:
                    cell.fill = source_fills.get(source, PatternFill())

            row_idx += 1

    # ── Column widths ──
    col_widths = {
        "A": 6,    # 序号
        "B": 22,   # 数据源
        "C": 12,   # 信号类型
        "D": 45,   # 标题
        "E": 60,   # 反馈内容
        "F": 24,   # 反馈时间
        "G": 12,   # 评分/热度
        "H": 14,   # 评分类型
        "I": 45,   # URL
        "J": 16,   # 作者
        "K": 30,   # 关键词
        "L": 12,   # 品类
        "M": 45,   # 附加信息
        "N": 20,   # 采集时间
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:N{row_idx - 1}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("数据统计")
    ws2.cell(row=1, column=1, value="数据源").font = Font(bold=True, size=12)
    ws2.cell(row=1, column=2, value="记录数").font = Font(bold=True, size=12)
    ws2.cell(row=1, column=3, value="说明").font = Font(bold=True, size=12)

    descriptions = {
        "amz_dataset": "Amazon真实产品评论，含1-5星评分和评论文本",
        "reddit": "Reddit用户讨论帖，含评论数，反映真实痛点",
        "github": "GitHub开源项目issue，含feature request和enhancement",
        "hackernews": "Hacker News热门帖子，含点赞数和评论数",
        "google_trends": "Google搜索趋势数据，含趋势值和相关搜索",
        "youtube": "YouTube选品相关视频，含播放量和频道信息",
    }

    r2 = 2
    total = 0
    for src in QUALITY_SOURCES:
        count = conn.execute("SELECT COUNT(*) FROM demand_signals WHERE source=?", (src,)).fetchone()[0]
        ws2.cell(row=r2, column=1, value=SOURCE_CN.get(src, src))
        ws2.cell(row=r2, column=2, value=count)
        ws2.cell(row=r2, column=3, value=descriptions.get(src, ""))
        total += count
        r2 += 1

    ws2.cell(row=r2, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=r2, column=2, value=total).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 55

    # Signal type stats
    r2 += 3
    ws2.cell(row=r2, column=1, value="信号类型").font = Font(bold=True, size=12)
    ws2.cell(row=r2, column=2, value="记录数").font = Font(bold=True, size=12)
    r2 += 1
    for row in conn.execute("""
        SELECT signal_type, COUNT(*) as cnt 
        FROM demand_signals 
        WHERE source IN ({})
        GROUP BY signal_type ORDER BY cnt DESC
    """.format(",".join(f"'{s}'" for s in QUALITY_SOURCES))):
        ws2.cell(row=r2, column=1, value=SIGNAL_TYPE_CN.get(row["signal_type"], row["signal_type"]))
        ws2.cell(row=r2, column=2, value=row["cnt"])
        r2 += 1

    wb.save(OUTPUT_PATH)
    conn.close()

    print(f"Exported {total} records to {OUTPUT_PATH}")
    print(f"Sheets: '质量反馈数据' ({total} rows), '数据统计'")


if __name__ == "__main__":
    main()

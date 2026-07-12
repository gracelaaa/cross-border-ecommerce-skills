#!/usr/bin/env python3
"""
Generate 2026 Cross-border E-commerce Opportunity Analysis Report (Excel).
Based on 760 quality demand signals + 2026 market research data.
"""
import sqlite3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/grace/WorkBuddy/2026-07-07-21-28-00/cross-border-ecommerce-skills/pipeline/demand_signals.db"
OUTPUT = "/Users/grace/WorkBuddy/2026-07-07-21-28-00/2026_opportunity_analysis.xlsx"

wb = Workbook()

# ── Styles ──
thin_border = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
title_font = Font(name="Microsoft YaHei", bold=True, size=16, color="FFFFFF")
title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
section_font = Font(name="Microsoft YaHei", bold=True, size=13, color="FFFFFF")
section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
body_font = Font(name="Microsoft YaHei", size=10)
body_bold = Font(name="Microsoft YaHei", size=10, bold=True)
body_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

region_fills = {
    "美国": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "欧盟": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "东南亚": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "拉美": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "印度": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "中东": PatternFill(start_color="D6F0E8", end_color="D6F0E8", fill_type="solid"),
}

# ════════════════════════════════════════════════════════
# Sheet 1: 机会总览
# ════════════════════════════════════════════════════════
ws = wb.active
ws.title = "机会总览"

# Title
ws.merge_cells("A1:G1")
c = ws.cell(row=1, column=1, value="2026 出海选品机会分析报告")
c.font = title_font
c.fill = title_fill
c.alignment = center_align
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:G2")
c = ws.cell(row=2, column=1, value="基于 760 条需求信号数据 + 2026 全球跨境电商市场数据 | 数据源：Amazon评论/Reddit/Google Trends/YouTube/HackerNews")
c.font = Font(name="Microsoft YaHei", size=10, italic=True, color="666666")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# Headers
headers = ["市场", "2026增速", "市场规模", "数据信号数", "机会品类数", "最高优先级品类", "主要风险"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border
ws.row_dimensions[4].height = 28

data = [
    ["美国 US", "8%", "$1.46万亿", "234条", "4", "高质量美容工具\n(差评集中)", "800免税取消\n关税+15%"],
    ["欧盟 EU", "4.6%", "$1.28万亿", "130+条", "4", "EU制造安全/\n隐私产品", "DST税3%\n佣金+0.5-1%"],
    ["东南亚 SEA", "18-23%", "高速增长", "间接信号", "4", "轻量AI\n电商SaaS", "竞争白热化\n广告费+25%"],
    ["拉美 LatAm", "22%", "Mercado Libre", "间接信号", "4", "电子配件\n(增速35%)", "汇率波动\n物流基建弱"],
    ["印度 India", "17.6%", "高速增长", "少量信号", "2", "电子产品\n验真服务", "数据本地化\nSNI认证"],
    ["中东 MEA", "高消费力", "Noon/Amazon.sa", "无直接信号", "2", "高端美妆\n个护", "市场数据少\n需本地化"],
]

for r, row_data in enumerate(data, 5):
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills.get(val.split()[0], PatternFill())
    ws.row_dimensions[r].height = 50

col_widths = {"A": 14, "B": 10, "C": 14, "D": 12, "E": 10, "F": 18, "G": 16}
for letter, width in col_widths.items():
    ws.column_dimensions[letter].width = width


# ════════════════════════════════════════════════════════
# Sheet 2: 美国市场
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("美国市场")

ws2.merge_cells("A1:F1")
c = ws2.cell(row=1, column=1, value="美国市场 US — 机会分析")
c.font = section_font
c.fill = section_fill
c.alignment = center_align
ws2.row_dimensions[1].height = 30

# Market context
ws2.merge_cells("A3:F3")
c = ws2.cell(row=3, column=1, value="市场背景：2026年规模$1.46万亿，增速8%。800美元免税取消，直邮模式受冲击，Temu订单-18%。需转海外仓备货，客单价建议$30+。")
c.font = Font(name="Microsoft YaHei", size=10, color="C00000")
c.alignment = body_align

# Data signals
ws2.merge_cells("A5:F5")
c = ws2.cell(row=5, column=1, value="数据信号来源（234条）")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")

signal_headers = ["数据源", "信号类型", "条数", "关键发现", "信号强度", "来源"]
for col, h in enumerate(signal_headers, 1):
    c = ws2.cell(row=6, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

us_signals = [
    ["Amazon评论数据集", "差评(pain_point)", "113", "冰滚轮193条/护肤117条/化妆刷55条中质量痛点集中", "★★★★★", "1-2星评论"],
    ["Amazon评论数据集", "好评(review)", "371", "love(94)/price(85)/recommend(68)/soft(34)", "★★★★", "4-5星评论"],
    ["Reddit", "痛点/讨论", "20+", "Whisker宠物科技(120评)/Arcteryx退货(100评)/Poshmark转售(110评)", "★★★★", "高评论数帖"],
    ["YouTube", "视频内容", "37", "数字产品3视频26万播放/选品教程86万总播放", "★★★", "高播放量"],
    ["Google Trends", "趋势", "20+", "amazon fba step by step(200,breakout)/what to sell(100)", "★★★★", "趋势值"],
]
for r, row_data in enumerate(us_signals, 7):
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
    ws2.row_dimensions[r].height = 40

# Opportunities
ws2.merge_cells("A14:F14")
c = ws2.cell(row=14, column=1, value="机会品类详解")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")

opp_headers = ["机会品类", "痛点/需求证据", "目标客群", "建议客单价", "竞争格局", "切入策略"]
for col, h in enumerate(opp_headers, 1):
    c = ws2.cell(row=15, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

us_opps = [
    [
        "高质量美容工具\n(冰滚轮/化妆刷/剃毛器)",
        "Amazon 484条评论中冰滚轮193条、化妆刷55条。\n差评关键词：cheap(5)/broke(3)/flimsy(2)/smell(7)。\n典型差评：'WASTE OF MONEY! They Melt & Snap Apart'、\n'Very weak and poorly made- broke in half the 2nd time'",
        "25-45岁女性\n个人护理消费者",
        "$15-35\n(海外仓发货)",
        "现有产品质量普遍差\n差评率极高\n品牌空白",
        "主打材质升级(316不锈钢/\n医疗级硅胶)，强调耐用性\n和安全性，用差评对比图\n做详情页"
    ],
    [
        "宠物科技替代品\n(智能猫砂盆/喂食器)",
        "Reddit: 'I'll never purchase another Whisker product'\n(120评论) — Whisker/Litter-Robot品牌信任危机\n用户抱怨产品质量和售后服务",
        "养猫/养狗家庭\n年收入$50K+",
        "$80-200\n(高客单价分摊关税)",
        "Whisker品牌信任崩塌\n市场出现替代窗口",
        "定位'可靠平替'，强调\n售后保障和性价比。\n通过Reddit/TikTok种草\n收割Whisker流失用户"
    ],
    [
        "数字产品/SaaS\n(模板/课程/工具)",
        "YouTube: '3 TINY Digital Products Guaranteed to Sell in 2026'\n(13万播放)/'5 Best Digital Products'(5.7万播放)\nGoogle Trends: 'ai dropshipping'(88)趋势上升",
        "电商卖家/\n内容创作者",
        "$9-49\n(数字交付零关税)",
        "蓝海市场\n2026年趋势爆发",
        "零关税零物流成本！\n做选品模板/店铺装修/\nAI选品工具。通过\nYouTube/Gumroad销售"
    ],
    [
        "户外装备配件\n(修复/替换零件)",
        "Reddit: 'I was denied a return for my Arcteryx coming apart'\n(100评论) — 高端户外装备售后痛点\n用户寻求可自行修复的配件",
        "户外运动爱好者\n中高收入",
        "$20-50\n(配件定位)",
        "大品牌不卖配件\n第三方空白",
        "做Arcteryx/Patagonia等\n品牌的兼容修复配件。\n避开品牌侵权，定位\n'universal repair kit'"
    ],
]
for r, row_data in enumerate(us_opps, 16):
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["美国"]
    ws2.row_dimensions[r].height = 90

col_widths2 = {"A": 18, "B": 35, "C": 16, "D": 14, "E": 16, "F": 28}
for letter, width in col_widths2.items():
    ws2.column_dimensions[letter].width = width


# ════════════════════════════════════════════════════════
# Sheet 3: 欧盟市场
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("欧盟市场")

ws3.merge_cells("A1:F1")
c = ws3.cell(row=1, column=1, value="欧盟市场 EU — 机会分析")
c.font = section_font
c.fill = section_fill
c.alignment = center_align
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A3:F3")
c = ws3.cell(row=3, column=1, value="市场背景：2026年规模$1.28万亿，增速4.6%。DST税3%生效，平台佣金+0.5-1%。AliExpress EU价格翻倍，BuyFromEU运动兴起。")
c.font = Font(name="Microsoft YaHei", size=10, color="C00000")
c.alignment = body_align

ws3.merge_cells("A5:F5")
c = ws3.cell(row=5, column=1, value="数据信号来源（130+条）")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="3B6D11")

for col, h in enumerate(signal_headers, 1):
    c = ws3.cell(row=6, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

eu_signals = [
    ["Reddit", "痛点/讨论", "3", "r/BuyFromEU: 'Alternative EU VPN product to Proton'(70评)\n欧洲消费者主动寻找美国品牌替代品", "★★★★★", "高评论数"],
    ["Reddit", "痛点", "2", "r/Aliexpress: 'New EU fee coming'(130评)+\n'prices doubled for EU'(50评) — AliExpress涨价后空白", "★★★★★", "高评论数"],
    ["HackerNews", "讨论", "2", "Top researchers leave USA for Netherlands(199赞)\nEurope's websites served by US vendors(134赞)", "★★★", "高赞数"],
    ["Amazon评论", "评论", "484", "好评关键词: natural(13)/quality(25) — 欧洲消费者\n重视天然成分和品质", "★★★★", "评论数据"],
]
for r, row_data in enumerate(eu_signals, 7):
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
    ws3.row_dimensions[r].height = 40

ws3.merge_cells("A12:F12")
c = ws3.cell(row=12, column=1, value="机会品类详解")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="3B6D11")

for col, h in enumerate(opp_headers, 1):
    c = ws3.cell(row=13, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

eu_opps = [
    [
        "EU制造安全/隐私产品\n(VPN/数据安全/加密通信)",
        "Reddit r/BuyFromEU: 用户主动求替代Proton VPN的\nEU产品(70评论)。BuyFromEU运动正在扩大。\nHN: 欧洲数字主权讨论热度高",
        "注重隐私的\n欧洲消费者",
        "€5-15/月\n订阅制",
        "Proton(Swiss)主导\n但用户不满\n替代品稀缺",
        "定位'100% EU-made'\n强调GDPR合规和数据\n本地存储。通过\nBuyFromEU社区推广"
    ],
    [
        "天然成分美妆个护\n(有机/纯素/无动物实验)",
        "Amazon好评关键词: natural(13)、quality(25)\n欧洲消费者重视天然成分和品质\nDST税后中国低价美妆竞争力下降",
        "25-40岁女性\n环保意识强",
        "€15-40",
        "低价中国品牌受DST冲击\nEU本土品牌溢价高\n中间价格带空白",
        "主打EU合规认证(ECOCERT/\nCOSMOS)，价格定位中端\n(€15-40)，填补AliExpress\n涨价后的价格空白"
    ],
    [
        "中高端日用品直供\n(家居清洁/厨房工具)",
        "AliExpress EU价格翻倍后，$5-15日用品\n出现价格真空。Reddit用户抱怨'prices\ndoubled or tripled for EU'",
        "欧洲家庭\n中等收入",
        "€8-25",
        "AliExpress涨价\nTemu受关税影响\n本地超市价格高",
        "通过海外仓直发模式\n填补$8-25价格带。\n重点品类：厨房工具、\n收纳用品、清洁工具"
    ],
    [
        "可持续时尚\n(环保材质/二手/修复)",
        "HN: 欧洲数字主权和可持续性讨论活跃\nReddit: Etsy卖家不满(70评论)暗示手工艺品\n市场存在供需错配",
        "20-35岁\n环保意识消费者",
        "€20-60",
        "SHEIN受DST和环保争议冲击\nEtsy卖家流失\n可持续品牌溢价过高",
        "做可持续材质(有机棉/\n再生纤维)的基础款。\n通过Etsy/独立站销售\n主打'EU-made'标签"
    ],
]
for r, row_data in enumerate(eu_opps, 14):
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["欧盟"]
    ws3.row_dimensions[r].height = 90

for letter, width in col_widths2.items():
    ws3.column_dimensions[letter].width = width


# ════════════════════════════════════════════════════════
# Sheet 4: 东南亚 & 拉美 & 印度 & 中东 (合并)
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet("新兴市场")

ws4.merge_cells("A1:F1")
c = ws4.cell(row=1, column=1, value="新兴市场 — 东南亚 / 拉美 / 印度 / 中东")
c.font = section_font
c.fill = section_fill
c.alignment = center_align
ws4.row_dimensions[1].height = 30

# SEA
ws4.merge_cells("A3:F3")
c = ws4.cell(row=3, column=1, value="东南亚 SEA | 增速18-23% | TikTok Shop +150% | 印尼+22% 泰国+20.6% 菲律宾+23%")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="854F0B")
c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
c.alignment = body_align

sea_opps = [
    ["轻量AI电商SaaS\n(选品/客服/广告工具)", "HN: 'Small AI Models Gain Traction in unreliable networks'(217赞)\nYouTube: AI Dropshipping课程68K播放\nGoogle Trends: 'ai dropshipping'(88)", "东南亚电商卖家", "$5-29/月", "蓝海\n本地工具稀缺", "做轻量级Web应用\n支持离线/弱网使用\n面向Shopee/TikTok卖家"],
    ["美妆个护", "马来西亚40%订单来自跨境，美妆最受欢迎\n印尼/泰国年轻人口结构", "18-30岁女性", "$3-12", "韩国品牌主导高端\n中国品牌主导低端\n中间空白", "定位K-beauty平替\n价格$5-12\n通过TikTok Shop销售"],
    ["平价时尚", "东南亚年轻人口多，社交媒体驱动消费\nShopee广告费+25%但体量仍大", "18-25岁", "$5-15", "SHEIN/Temu覆盖\n但本地化不足", "做东南亚尺码/审美\n本地化款。通过\nShopee+TikTok双平台"],
    ["TikTok内容电商\n(直播带货服务)", "TikTok Shop GMV +150%\n菲律宾23%增速由社交电商驱动", "中小卖家", "服务费\n$200-500/月", "直播MCN稀缺\n服务商少", "做TikTok Shop代运营/\n直播培训/短视频制作\n服务商定位"],
]
ws4.merge_cells("A5:F5")
c = ws4.cell(row=5, column=1, value="机会品类")
c.font = Font(name="Microsoft YaHei", bold=True, size=11, color="854F0B")

for col, h in enumerate(opp_headers, 1):
    c = ws4.cell(row=6, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

for r, row_data in enumerate(sea_opps, 7):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["东南亚"]
    ws4.row_dimensions[r].height = 80

# LatAm
start = 12
ws4.merge_cells(f"A{start}:F{start}")
c = ws4.cell(row=start, column=1, value="拉美 LatAm | 增速22% | Mercado Libre | 巴西+墨西哥为主 | TikTok Shop Q2开放巴西")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="993C1D")
c.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
c.alignment = body_align

ws4.merge_cells(f"A{start+2}:F{start+2}")
c = ws4.cell(row=start+2, column=1, value="机会品类")
c.font = Font(name="Microsoft YaHei", bold=True, size=11, color="993C1D")

for col, h in enumerate(opp_headers, 1):
    c = ws4.cell(row=start+3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

latam_opps = [
    ["电子配件\n(手机壳/充电器/数据线)", "市场数据: 电子配件增速35%\n巴西站客单价$35+复购率40%\nReddit: 卖家讨论Mercado Libre机会", "18-35岁\n智能手机用户", "$5-20", "本地供应不足\n中国品牌品质参差", "做设计感手机壳+\n快充配件组合\n通过Mercado Libre销售"],
    ["家居装饰\n(灯具/墙贴/收纳)", "市场数据: 家居装饰增速28%\n拉美家庭装饰需求旺盛", "25-40岁家庭", "$8-25", "本地选择少\n宜家覆盖不足", "做热带风格/彩色系\n家居装饰\n适应拉美审美"],
    ["运动户外\n(瑜伽垫/跑步装备)", "市场数据: 运动户外增速25%\n健康生活方式趋势", "20-35岁\n健身人群", "$10-30", "国际品牌溢价高\n本地品牌品质差", "做中端品质运动装备\n$10-30价格带\n填补品牌空白"],
    ["平价美妆", "拉美女性美妆消费高\nMercado Libre美妆类目增长快", "18-30岁女性", "$3-12", "本地品牌低端\n国际品牌昂贵", "K-beauty风格平价美妆\n通过TikTok Shop巴西站\n内容种草"],
]
for r, row_data in enumerate(latam_opps, start+4):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["拉美"]
    ws4.row_dimensions[r].height = 80

# India
start2 = start + 9
ws4.merge_cells(f"A{start2}:F{start2}")
c = ws4.cell(row=start2, column=1, value="印度 India | 增速17.6% | Amazon/Flipkart | 数据本地化法规")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="534AB7")
c.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
c.alignment = body_align

ws4.merge_cells(f"A{start2+2}:F{start2+2}")
c = ws4.cell(row=start2+2, column=1, value="机会品类")
c.font = Font(name="Microsoft YaHei", bold=True, size=11, color="534AB7")

for col, h in enumerate(opp_headers, 1):
    c = ws4.cell(row=start2+3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

india_opps = [
    ["电子产品验真服务", "Reddit r/LaptopDealsIndia: 'Amazon/Flipkart laptops\nare used ones?' — 印度消费者对电商电子产品\n真伪存疑，信任缺口大", "网购电子产品\n消费者", "服务费\n₹99-299", "空白市场\n无成熟验真服务", "做到府验真/视频验真\n服务。与Flipkart/\nAmazon卖家合作"],
    ["平价电子配件", "印度智能手机普及率高\n配件需求大但品质信任低", "18-30岁", "₹99-499", "本地品质差\n中国品牌渠道弱", "做品质保证的电子配件\n强调 warranty 和\nquality check"],
]
for r, row_data in enumerate(india_opps, start2+4):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["印度"]
    ws4.row_dimensions[r].height = 80

# MEA
start3 = start2 + 7
ws4.merge_cells(f"A{start3}:F{start3}")
c = ws4.cell(row=start3, column=1, value="中东 MEA | 高消费力 | Noon/Amazon.sa | 政府数字化转型")
c.font = Font(name="Microsoft YaHei", bold=True, size=12, color="0F6E56")
c.fill = PatternFill(start_color="D6F0E8", end_color="D6F0E8", fill_type="solid")
c.alignment = body_align

ws4.merge_cells(f"A{start3+2}:F{start3+2}")
c = ws4.cell(row=start3+2, column=1, value="机会品类（注：本市场无直接数据信号，基于市场研究）")
c.font = Font(name="Microsoft YaHei", bold=True, size=11, color="0F6E56")

for col, h in enumerate(opp_headers, 1):
    c = ws4.cell(row=start3+3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

mea_opps = [
    ["高端美妆个护", "沙特/阿联酋人均美妆消费高\n宗教文化偏好halal认证产品", "25-45岁女性\n高收入", "$30-80", "国际品牌主导\n但halal认证产品稀缺", "做halal认证美妆\n通过Noon和Amazon.sa\n强调中东本地化"],
    ["智能家居电子", "政府数字化转型政策\n智能家居需求增长", "28-45岁\n科技爱好者", "$25-100", "国际品牌贵\n本地品牌少", "做阿拉伯语智能音箱/\n智能家居套装\n通过Noon销售"],
]
for r, row_data in enumerate(mea_opps, start3+4):
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 1:
            c.font = body_bold
            c.fill = region_fills["中东"]
    ws4.row_dimensions[r].height = 80

for letter, width in col_widths2.items():
    ws4.column_dimensions[letter].width = width


# ════════════════════════════════════════════════════════
# Sheet 5: 数据源信号明细
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet("数据信号明细")

ws5.merge_cells("A1:E1")
c = ws5.cell(row=1, column=1, value="数据信号明细 — 各源关键发现")
c.font = section_font
c.fill = section_fill
c.alignment = center_align
ws5.row_dimensions[1].height = 30

detail_headers = ["数据源", "条数", "关键发现", "机会指向", "信号强度"]
for col, h in enumerate(detail_headers, 1):
    c = ws5.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

details = [
    ["Amazon评论数据集\n(All_Beauty)", "484", "产品类型分布：冰滚轮193条/护肤117条/化妆刷55条/\n指甲工具24条/剃毛器8条\n差评TOP词：hair(19)/smell(7)/razor(6)/cheap(5)\n好评TOP词：great(94)/love(58)/color(45)/easy(42)", "美国市场美容工具质量升级\n天然成分护肤品牌机会", "★★★★★"],
    ["Reddit", "73", "EU痛点：AliExpress涨价(130+50评)/BuyFromEU(70评)\n美国痛点：Whisker宠物(120评)/Arcteryx退货(100评)\n需求信号：'where can I buy'多帖(travel router/DS games/\nface paint/red currant)", "EU本地化替代品\n宠物科技替代\n户外装备配件", "★★★★★"],
    ["Google Trends", "47", "'amazon fba step by step'(200,breakout)\n'what to sell on amazon to make money'(100)\n'ai dropshipping'(88)上升趋势\n'shopify 登录'(60)中国卖家涌入", "AI电商工具需求\n卖家选品工具需求\n数字产品趋势", "★★★★"],
    ["YouTube", "37\n(86万播放)", "数字产品3视频26万播放(2026趋势)\nAI Dropshipping课程68K播放\n选品教程系列总播放86万\n月度TOP 10产品系列持续更新", "数字产品/SaaS\nAI电商工具\n选品内容创作", "★★★"],
    ["HackerNews", "59", "边缘AI在弱网地区受关注(217赞)\n欧洲数字主权讨论(199+134赞)\n开源硬件路由器需求(744赞)", "东南亚轻量AI工具\nEU安全/隐私产品\n开源硬件配件", "★★★"],
    ["GitHub Issues", "60", "Amazon Seller API问题(114赞)\n各类工具feature request\nMCP/AI工具生态活跃", "Amazon卖家工具开发\nAI电商工具生态", "★★"],
]

for r, row_data in enumerate(details, 4):
    for col, val in enumerate(row_data, 1):
        c = ws5.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
    ws5.row_dimensions[r].height = 80

col_widths5 = {"A": 18, "B": 10, "C": 40, "D": 22, "E": 10}
for letter, width in col_widths5.items():
    ws5.column_dimensions[letter].width = width


# ════════════════════════════════════════════════════════
# Sheet 6: 优先级矩阵
# ════════════════════════════════════════════════════════
ws6 = wb.create_sheet("优先级矩阵")

ws6.merge_cells("A1:F1")
c = ws6.cell(row=1, column=1, value="机会优先级矩阵 — 按 可行性 × 信号强度 × 利润空间 排序")
c.font = section_font
c.fill = section_fill
c.alignment = center_align
ws6.row_dimensions[1].height = 30

matrix_headers = ["排名", "市场", "机会品类", "数据信号强度", "可行性", "综合推荐"]
for col, h in enumerate(matrix_headers, 1):
    c = ws6.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

matrix = [
    [1, "美国", "高质量美容工具(冰滚轮/化妆刷)", "★★★★★ 484条评论直接证据", "★★★★★ 供应链成熟(义乌/1688)", "立即行动\n差评对比营销"],
    [2, "美国", "数字产品/SaaS(选品模板/工具)", "★★★★ YouTube 86万播放", "★★★★★ 零关税零物流", "立即行动\n最快变现"],
    [3, "欧盟", "EU制造安全/隐私产品", "★★★★★ Reddit 70评主动需求", "★★★ 需技术能力", "Q3启动\nBuyFromEU社区"],
    [4, "拉美", "电子配件(手机壳/充电器)", "★★★★ 增速35%+复购40%", "★★★★ 供应链成熟", "Q3启动\nMercado Libre"],
    [5, "美国", "宠物科技替代品(智能猫砂盆)", "★★★★ Reddit 120评信任危机", "★★★ 需硬件开发", "Q4规划\n收割Whisker流失"],
    [6, "东南亚", "轻量AI电商SaaS(选品/客服)", "★★★ HN 217赞+Trends 88", "★★★★ 纯软件", "Q4规划\n面向Shopee卖家"],
    [7, "欧盟", "天然成分美妆个护", "★★★★ Amazon好评natural(13)", "★★★ 需EU合规认证", "Q4规划\nECOCERT认证"],
    [8, "拉美", "家居装饰(灯具/墙贴)", "★★★ 增速28%", "★★★★ 轻物流", "2027规划\n本地化设计"],
    [9, "东南亚", "TikTok内容电商服务", "★★★ TikTok Shop +150%", "★★★ 需内容团队", "2027规划\n服务商定位"],
    [10, "印度", "电子产品验真服务", "★★ Reddit讨论信任问题", "★★ 需本地运营", "观察期\n等待数据积累"],
]

for r, row_data in enumerate(matrix, 4):
    for col, val in enumerate(row_data, 1):
        c = ws6.cell(row=r, column=col, value=val)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if col == 2:
            c.font = body_bold
            c.fill = region_fills.get(val, PatternFill())
        if col == 1:
            c.font = body_bold
            c.alignment = center_align
    ws6.row_dimensions[r].height = 50

col_widths6 = {"A": 6, "B": 10, "C": 24, "D": 22, "E": 18, "F": 18}
for letter, width in col_widths6.items():
    ws6.column_dimensions[letter].width = width


# Save
wb.save(OUTPUT)
print(f"Report saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
